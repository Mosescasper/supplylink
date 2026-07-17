"""
import_quickbooks_items.py
SupplyLink -- Bulk import of a QuickBooks item export (.xlsx) into the Item table.

Only rows with Type == "Stock Part" are imported (real physical inventory --
drugs and supplies). "Service", "Subtotal", "Group", "VAT Item", etc. rows
are billing/fee entries, not stock items, and are skipped.

Note: every named sheet in this workbook (pharmacy, nutrition, dental, etc.)
contains the SAME full item list -- the tab names don't actually filter by
department. This script defaults to reading "all QB  items" (the master
sheet) so you only import each item once. If you later get a file that
genuinely differs per sheet, pass --sheet to target a specific one.

Usage:
    python import_quickbooks_items.py "quick_books_items.xlsx" --dry-run
    python import_quickbooks_items.py "quick_books_items.xlsx"

Run with --dry-run first to preview counts without writing to the database.
"""

import argparse

import openpyxl

from app import app
from extensions import db
from models import Item

# Column positions in the QuickBooks export (0-indexed, values_only row tuple)
COL_TYPE = 2
COL_ITEM_CODE = 3
COL_DESCRIPTION = 4
COL_QUANTITY_ON_HAND = 12
COL_COST = 13
COL_PRICE = 16
COL_REORDER_PT = 19


def safe_float(value, default=0.0):
    """Coerce a cell value to float, tolerating blanks/None."""
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def import_items(xlsx_path, sheet_name="all QB  items", dry_run=False):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        candidates = [s for s in wb.sheetnames if sheet_name.strip() in s.strip()]
        if not candidates:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
            )
        sheet_name = candidates[0]

    ws = wb[sheet_name]
    print(f"Reading sheet: '{sheet_name}'")

    created, updated, skipped_non_stock, skipped_missing = 0, 0, 0, 0
    truncated_names = []

    with app.app_context():
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            item_type = row[COL_TYPE]

            if item_type != "Stock Part":
                skipped_non_stock += 1
                continue

            code = row[COL_ITEM_CODE]
            description = row[COL_DESCRIPTION]

            if not code or not description:
                skipped_missing += 1
                continue

            sku = str(code).strip()[:50]

            name = str(description).strip()
            if len(name) > 200:
                truncated_names.append((sku, name))
                name = name[:200]

            cost = safe_float(row[COL_COST])
            price = safe_float(row[COL_PRICE])
            # QuickBooks sometimes has Cost=0 with the real price only in
            # the Price column (or vice versa) -- prefer Cost, fall back
            # to Price so we don't import a batch of zero-cost items.
            unit_cost = cost if cost > 0 else price

            reorder_level = safe_float(row[COL_REORDER_PT], default=0)

            item = Item.query.filter_by(sku=sku).first()
            if item is None:
                item = Item(
                    sku=sku,
                    name=name,
                    unit_of_issue="Units",
                    unit_cost=unit_cost,
                    reorder_level=reorder_level,
                )
                if not dry_run:
                    db.session.add(item)
                created += 1
            else:
                item.name = name
                item.unit_cost = unit_cost
                if reorder_level:
                    item.reorder_level = reorder_level
                updated += 1

        if dry_run:
            db.session.rollback()
            print("\n-- DRY RUN: nothing was written to the database --")
        else:
            db.session.commit()

        print(f"\nItems created:  {created}")
        print(f"Items updated:  {updated}")
        print(f"Skipped (not a Stock Part row): {skipped_non_stock}")
        print(f"Skipped (missing code/description): {skipped_missing}")

        if truncated_names:
            print(
                f"\n{len(truncated_names)} item names were longer than 200 characters "
                f"and were truncated to fit the database column:"
            )
            for sku, name in truncated_names[:10]:
                print(f"  - {sku}: {name[:80]}...")
            if len(truncated_names) > 10:
                print(f"  ... and {len(truncated_names) - 10} more")

        print(
            "\nNote: 'Reorder Pt (Min)' was blank for every row in this file, "
            "so all imported items have reorder_level = 0. Set real reorder "
            "levels per item afterwards (via the Edit Item form, or a follow-up "
            "bulk update) once you know realistic thresholds."
        )
        print(
            "Also note: this QuickBooks export mixes drugs with non-drug "
            "hospital supplies (theatre consumables, orthopaedic hardware, "
            "nutrition products, etc.) since QuickBooks doesn't separate them "
            "by category. Assign Category values afterwards if you want to "
            "filter by 'drugs only' in SupplyLink."
        )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Import a QuickBooks item export into SupplyLink's Item table"
    )
    parser.add_argument("xlsx_path", help="Path to the QuickBooks .xlsx export")
    parser.add_argument(
        "--sheet", default="all QB  items",
        help="Sheet name to import (default: 'all QB  items' -- the master list)"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview counts without writing to the database"
    )
    args = parser.parse_args()

    import_items(args.xlsx_path, args.sheet, dry_run=args.dry_run)