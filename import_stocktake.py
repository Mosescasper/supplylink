"""
Import script for PHARMACY_STORE_STOCKTAKE__REPORT_JUNE__2026.xlsx
into SupplyLink's Drug Store.

USAGE (run from inside your project folder, e.g. supplylink-clean):

    python import_stocktake.py "path\\to\\PHARMACY_STORE_STOCKTAKE__REPORT_JUNE__2026__1_.xlsx"

WHAT IT DOES:
  - Reads the "JUNE STOCK TAKE" sheet (headers on row 6, data from row 7)
  - For each row: KNH CAT -> Item.sku, ITEM DESCRIPTION -> Item.name,
    UNIT OF ISSUE -> Item.unit_of_issue, UNIT PRICE -> Item.unit_cost
  - Creates the item if it doesn't exist yet, or updates name/cost/unit
    if it does (matched by sku)
  - Creates or updates a single Batch per item at location "Drug Store",
    with quantity_received/quantity_remaining set from PHYSICAL STOCK
  - Parses REMARKS (e.g. "6/27", "1/2028") into an expiry_date (last day
    of that month). Blank/unparseable remarks default to 2 years from
    today, and are flagged in the summary so you can check them manually.
  - Duplicate SKUs in the sheet are merged: physical stock quantities are
    SUMMED into a single item/batch rather than creating conflicting
    duplicate records.
  - Uses a fixed batch_number per item (STOCKTAKE-2026-06-<sku>), so
    re-running this script is safe -- it updates the same batch instead
    of creating new ones each time.
  - Nothing is committed to the database until the whole file has been
    read and validated -- if something goes wrong partway through, no
    partial data is written.

This script does NOT touch StockMovement / audit-trail records deliberately
-- it's meant for an initial "here is what we physically counted" load, not
a live receipt/transfer transaction. If you want stock movements logged for
this import too, let me know and I'll add that as a separate, explicit step.
"""

import sys
import calendar
from datetime import date, timedelta
from collections import defaultdict

import openpyxl

# Pull in your actual app/db/models so this script uses the exact same
# validation and schema your app already relies on.
from app import app
from extensions import db
from models import Item, Batch

SHEET_NAME_CONTAINS = "JUNE"
LOCATION = "Drug Store"
HEADER_ROW = 6
DATA_START_ROW = 7

# Column indices (0-based, matching openpyxl's row tuples from this file)
COL_SKU = 1
COL_NAME = 2
COL_DOSAGE_FORM = 3
COL_UNIT_OF_ISSUE = 4
COL_UNIT_PRICE = 5
COL_PHYSICAL_STOCK = 7
COL_REMARKS = 13


def parse_expiry(remark, fallback_years_ahead=2):
    """Parse a REMARKS value like '6/27' or '1/2028' into a date (last day
    of that month). Returns (expiry_date, was_guessed) -- was_guessed is
    True when we had to fall back to a default, so the caller can flag it."""
    if not remark or not isinstance(remark, str) or "/" not in remark:
        fallback = date.today() + timedelta(days=365 * fallback_years_ahead)
        return fallback, True

    try:
        month_str, year_str = remark.strip().split("/")
        month = int(month_str)
        year = int(year_str)
        if year < 100:
            year += 2000
        last_day = calendar.monthrange(year, month)[1]
        return date(year, month, last_day), False
    except (ValueError, IndexError):
        fallback = date.today() + timedelta(days=365 * fallback_years_ahead)
        return fallback, True


def load_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = next(
        (name for name in wb.sheetnames if SHEET_NAME_CONTAINS in name.upper()),
        None,
    )
    if not sheet_name:
        raise SystemExit(
            f"Couldn't find a sheet containing '{SHEET_NAME_CONTAINS}'. "
            f"Sheets found: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    raw_rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))

    rows = []
    for r in raw_rows:
        sku = r[COL_SKU]
        name = r[COL_NAME]
        if not sku or not name:
            continue  # skip blank/section rows
        rows.append(r)
    return rows


def main():
    if len(sys.argv) != 2:
        print("Usage: python import_stocktake.py <path_to_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    rows = load_rows(xlsx_path)
    print(f"Read {len(rows)} data rows from the spreadsheet.")

    # --- Merge duplicate SKUs before touching the database ---
    merged = {}
    for r in rows:
        sku = str(r[COL_SKU]).strip()
        name = str(r[COL_NAME]).strip()
        unit_of_issue = (r[COL_UNIT_OF_ISSUE] or "Units")
        unit_of_issue = str(unit_of_issue).strip() or "Units"
        unit_price = float(r[COL_UNIT_PRICE] or 0)
        physical_stock = float(r[COL_PHYSICAL_STOCK] or 0)
        remark = r[COL_REMARKS]

        if sku in merged:
            merged[sku]["physical_stock"] += physical_stock
            merged[sku]["duplicate_count"] += 1
        else:
            merged[sku] = {
                "name": name,
                "unit_of_issue": unit_of_issue,
                "unit_price": unit_price,
                "physical_stock": physical_stock,
                "remark": remark,
                "duplicate_count": 1,
            }

    duplicates = {sku: v for sku, v in merged.items() if v["duplicate_count"] > 1}
    if duplicates:
        print(f"\nMerged {len(duplicates)} duplicate SKU(s) (quantities summed):")
        for sku, v in duplicates.items():
            print(f"  {sku}: {v['duplicate_count']} rows -> "
                  f"combined physical stock {v['physical_stock']:g}")

    # --- Apply to the database ---
    created_items = 0
    updated_items = 0
    created_batches = 0
    updated_batches = 0
    guessed_expiry = []

    with app.app_context():
        for sku, v in merged.items():
            item = Item.query.filter_by(sku=sku).first()
            if item is None:
                item = Item(
                    sku=sku,
                    name=v["name"],
                    unit_of_issue=v["unit_of_issue"],
                    unit_cost=v["unit_price"],
                    reorder_level=0,
                )
                db.session.add(item)
                db.session.flush()  # assigns item.id
                created_items += 1
            else:
                item.name = v["name"]
                item.unit_of_issue = v["unit_of_issue"]
                item.unit_cost = v["unit_price"]
                updated_items += 1

            expiry_date, was_guessed = parse_expiry(v["remark"])
            if was_guessed:
                guessed_expiry.append(sku)

            batch_number = f"STOCKTAKE-2026-06-{sku}"
            batch = Batch.query.filter_by(
                item_id=item.id, batch_number=batch_number, location=LOCATION
            ).first()

            if batch is None:
                batch = Batch(
                    item_id=item.id,
                    batch_number=batch_number,
                    expiry_date=expiry_date,
                    quantity_received=v["physical_stock"],
                    quantity_remaining=v["physical_stock"],
                    location=LOCATION,
                )
                db.session.add(batch)
                created_batches += 1
            else:
                batch.expiry_date = expiry_date
                batch.quantity_received = v["physical_stock"]
                batch.quantity_remaining = v["physical_stock"]
                updated_batches += 1

        db.session.commit()

    print("\n--- Import complete ---")
    print(f"Items created:   {created_items}")
    print(f"Items updated:   {updated_items}")
    print(f"Batches created: {created_batches}")
    print(f"Batches updated: {updated_batches}")
    if guessed_expiry:
        print(f"\n{len(guessed_expiry)} item(s) had an unparseable/blank REMARKS "
              f"value and got a default 2-year expiry -- worth checking manually:")
        for sku in guessed_expiry:
            print(f"  {sku}")


if __name__ == "__main__":
    main()